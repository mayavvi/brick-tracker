"""Service for parsing tracker Excel files into TaskItem lists."""

from __future__ import annotations

import logging
import re
from calendar import monthrange
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.models import TaskItem, TrackerFileInfo

logger = logging.getLogger(__name__)

# Header keywords used to locate the detail-header row (row 9 typically)
_MAIN_HEADER_KEYWORDS = {"主程序名", "撰写人", "主程序撰写人"}
_TARGET_SHEETS = {"spec", "数据集", "tfls"}

# Column-name mappings (key = canonical name, value = possible header texts)
_COL_ALIASES: dict[str, set[str]] = {
    "main_program": {"主程序名"},
    "main_date": {"说明文件创建日期", "数据产生日期", "TFLs产生日期", "tfls产生日期"},
    "main_person": {"撰写人", "主程序撰写人"},
    "main_status": set(),  # positional: first "状态" column
    "qc_program": {"QC程序名", "qc程序名"},
    "qc_date": {"QC完成日期", "qc完成日期"},
    "qc_person": {"审阅人", "QC程序撰写人", "qc程序撰写人"},
    "qc_content": {"QC内容", "qc内容"},
    "qc_status": set(),  # positional: second "状态" column
    "ddl": {"DDL", "ddl"},
    "batch": {"Batch", "batch"},
    "comment": {"Comment", "comment"},
}


def _default_ddl() -> date:
    """Return the last day of the current month as a fallback DDL."""
    today = date.today()
    _, last_day = monthrange(today.year, today.month)
    return date(today.year, today.month, last_day)


def _safe_date(value: Any) -> date | None:
    """Convert a cell value to ``date``, tolerating various formats."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    # Try "31DEC2025" style
    m = re.match(r"(\d{1,2})([A-Za-z]{3})(\d{4})", text)
    if m:
        try:
            return datetime.strptime(text, "%d%b%Y").date()
        except ValueError:
            pass
    # Try ISO format
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return None


def _safe_str(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _detect_header_row(ws: Worksheet, max_scan: int = 15) -> int | None:
    """Find the row index containing the detail-level column headers."""
    for row_idx in range(1, min(max_scan + 1, ws.max_row + 1)):
        for cell in ws[row_idx]:
            if cell.value and str(cell.value).strip() in _MAIN_HEADER_KEYWORDS:
                return row_idx
    return None


def _build_column_map(ws: Worksheet, header_row: int) -> dict[str, int]:
    """Map canonical column names to 0-based column indices."""
    col_map: dict[str, int] = {}
    status_indices: list[int] = []

    for col_idx, cell in enumerate(ws[header_row]):
        text = str(cell.value).strip() if cell.value else ""
        if not text:
            continue

        if text == "状态":
            status_indices.append(col_idx)
            continue

        for canonical, aliases in _COL_ALIASES.items():
            if text in aliases and canonical not in col_map:
                col_map[canonical] = col_idx
                break

    # First "状态" -> main_status, second -> qc_status
    if len(status_indices) >= 1:
        col_map["main_status"] = status_indices[0]
    if len(status_indices) >= 2:
        col_map["qc_status"] = status_indices[1]

    return col_map


def _cell_val(row: tuple, col_map: dict[str, int], key: str) -> Any:
    idx = col_map.get(key)
    if idx is None or idx >= len(row):
        return None
    return row[idx].value


def _extract_global_ddl(ws: Worksheet) -> date | None:
    """Extract the global DDL date from the header area (typically row 7)."""
    for row_idx in range(1, min(10, ws.max_row + 1)):
        cell = ws.cell(row=row_idx, column=1)
        if cell.value and "DDL" in str(cell.value):
            text = str(cell.value)
            m = re.search(r"(\d{1,2}[A-Za-z]{3}\d{4})", text)
            if m:
                return _safe_date(m.group(1))
    return None


def _normalize_sheet_type(name: str) -> str:
    """Normalise sheet name to one of SPEC / 数据集 / TFLs."""
    lower = name.strip().lower()
    if lower == "spec":
        return "SPEC"
    if lower in ("数据集",):
        return "数据集"
    if lower == "tfls":
        return "TFLs"
    return name


def parse_tracker_file(info: TrackerFileInfo) -> list[TaskItem]:
    """Parse all data sheets in a single tracker Excel file."""
    path = Path(info.file_path)
    if not path.exists():
        logger.warning("Tracker file not found: %s", path)
        return []

    try:
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    except Exception:
        logger.exception("Failed to open workbook: %s", path)
        return []

    tasks: list[TaskItem] = []
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() not in _TARGET_SHEETS:
            continue
        ws = wb[sheet_name]
        sheet_tasks = _parse_sheet(ws, info, _normalize_sheet_type(sheet_name))
        tasks.extend(sheet_tasks)

    wb.close()
    return tasks


def _parse_sheet(
    ws: Worksheet,
    info: TrackerFileInfo,
    sheet_type: str,
) -> list[TaskItem]:
    """Parse a single sheet into a list of TaskItem."""
    header_row = _detect_header_row(ws)
    if header_row is None:
        logger.debug(
            "No header row found in sheet '%s' of %s", ws.title, info.file_name
        )
        return []

    col_map = _build_column_map(ws, header_row)
    if not col_map:
        return []

    global_ddl = _extract_global_ddl(ws)

    tasks: list[TaskItem] = []
    last_category: str | None = None

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        # Col A = category, Col B = item name
        cat_val = _safe_str(row[0].value) if len(row) > 0 else None
        name_val = _safe_str(row[1].value) if len(row) > 1 else None

        if cat_val:
            last_category = cat_val

        # Skip rows without an item name and without any main/qc person
        main_person = _safe_str(_cell_val(row, col_map, "main_person"))
        qc_person = _safe_str(_cell_val(row, col_map, "qc_person"))
        if not name_val and not main_person and not qc_person:
            continue

        row_ddl = _safe_date(_cell_val(row, col_map, "ddl"))
        effective_ddl = row_ddl or global_ddl or _default_ddl()

        tasks.append(
            TaskItem(
                study_id=info.study_id,
                compound=info.compound,
                task_purpose=info.task_purpose,
                sheet_type=sheet_type,
                category=last_category,
                item_name=name_val or "",
                main_program=_safe_str(_cell_val(row, col_map, "main_program")),
                main_date=_safe_date(_cell_val(row, col_map, "main_date")),
                main_person=main_person,
                main_status=_safe_str(_cell_val(row, col_map, "main_status")),
                qc_program=_safe_str(_cell_val(row, col_map, "qc_program")),
                qc_date=_safe_date(_cell_val(row, col_map, "qc_date")),
                qc_person=qc_person,
                qc_content=_safe_str(_cell_val(row, col_map, "qc_content")),
                qc_status=_safe_str(_cell_val(row, col_map, "qc_status")),
                ddl=effective_ddl,
                batch=_safe_str(_cell_val(row, col_map, "batch")),
                comment=_safe_str(_cell_val(row, col_map, "comment")),
            )
        )

    return tasks
